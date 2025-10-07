import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl.styles import Alignment, Border, Side


def main(file_path, money):
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path, data_only=True)  # 替换为您的文件名
    sheet1 = wb['Sheet1']

    # 创建Sheet2
    if 'Sheet2' in wb.sheetnames:
        wb.remove(wb['Sheet2'])
    sheet2 = wb.create_sheet('Sheet2')

    # 创建Sheet3
    if 'Sheet3' in wb.sheetnames:
        wb.remove(wb['Sheet3'])
    sheet3 = wb.create_sheet('Sheet3')

    # ==== 处理Sheet2 ====
    # 设置标题
    sheet2['A1'] = '号码'
    sheet2['B1'] = '职场'
    sheet2['C1'] = '总分钟数'
    sheet2['D1'] = '有效分钟数'

    # 收集每个号码的数据
    phone_list = []
    workplace_list = []

    # 遍历Sheet1的数据 (跳过标题行)
    for row in range(2, sheet1.max_row + 1):
        phone = sheet1[f'G{row}'].value  # 号码
        workplace = sheet1[f'H{row}'].value  # 职场
        phone_list.append(phone)
        workplace_list.append(workplace)

    # 去重
    workplace_list = [item for item in workplace_list if item not in ['#N/A', None]]
    phone_list_s = list(set(phone_list))
    workplace_list_s = list(set(workplace_list))

    # 写入Sheet2数据
    row_idx = 2
    for phone in phone_list_s:
        sheet2[f'A{row_idx}'] = phone
        sheet2[f'B{row_idx}'] = f'==VLOOKUP(A{row_idx},Sheet1!G:H,2,FALSE)'

        # 添加公式
        sheet2[f'C{row_idx}'] = f'=SUMIF(Sheet1!G:G, A{row_idx}, Sheet1!I:I)'
        sheet2[f'D{row_idx}'] = f'=MAX(C{row_idx}-400, 0)'
        row_idx += 1

    # ==== 处理Sheet3 ====
    # 设置标题
    sheet3['A1'] = '职场'
    sheet3['B1'] = '通话时长'
    sheet3['C1'] = '金额'

    # 设置标题样式（居中+边框）
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 应用标题样式
    for col in range(1, 4):
        cell = sheet3.cell(row=1, column=col)
        cell.alignment = center_alignment
        cell.border = thin_border

    # 写入Sheet3数据
    row_idx = 2
    m_idx = len(workplace_list_s) + 2
    for workplace in workplace_list_s:
        sheet3[f'A{row_idx}'] = workplace
        # 添加SUMIFS公式
        sheet3[f'B{row_idx}'] = f'==SUMIF(Sheet2!B:B,A{row_idx},Sheet2!D:D)'
        sheet3[f'C{row_idx}'] = f'==B{row_idx}/B{m_idx}*C{m_idx}'
        # 设置C列数字格式为保留1位小数
        sheet3[f'C{row_idx}'].number_format = '0.0'

        # 应用数据行样式
        for col in range(1, 4):
            cell = sheet3.cell(row=row_idx, column=col)
            cell.alignment = center_alignment
            cell.border = thin_border
        row_idx += 1

    # 合计行样式
    sheet3[f'A{row_idx}'] = '合计'
    sheet3[f'B{row_idx}'] = f'==SUM(B2:B{row_idx - 1})'
    sheet3[f'C{row_idx}'] = money
    for col in range(1, 4):
        cell = sheet3.cell(row=row_idx, column=col)
        cell.alignment = center_alignment
        cell.border = thin_border

    # 调整列宽
    for sheet in [sheet2, sheet3]:
        for col in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col)
            sheet.column_dimensions[col_letter].width = 15

    # 保存文件
    wb.save(file_path)
    print("处理完成！结果已保存到 {}".format(file_path))


# TODO 先取消密码保存好，然后修改文件位置和金额后运行程序
if __name__ == '__main__':
    file_path = r'C:\Users\jianghongyu\Desktop\广东金湾信息科技有限公司-云录音1(1).xlsx'
    main(file_path, 773.25)
